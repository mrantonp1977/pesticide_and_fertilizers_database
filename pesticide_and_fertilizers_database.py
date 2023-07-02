import tkinter.ttk
from tkinter import *
from PIL import ImageTk, Image
import sqlite3
import csv
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re


if __name__ == "__main__":

    root = Tk()
    root.title("Pesticide & Fertilizers Database")
    root.geometry("700x722")
    root.config(background="#CAE1DA")
    icon = PhotoImage(file="pesticide.logo.png")
    root.iconphoto(True, icon)

    conn = sqlite3.connect("pesticide_and_fertilizers_file.db")
    c = conn.cursor()


    title_label = Label(root, text=" ΦΥΤΟΦΑΡΜΑΚΑ & ΛΙΠΑΣΜΑΤΑ ", font=("arial black", 20), background="black",
                        relief="ridge", borderwidth=7, foreground="#34AD08")
    title_label.grid(row=0, columnspan=2, padx=80, pady=15)


    '''
    c.execute("""CREATE TABLE products (
                date integer,
                number integer,
                items integer,
                description text,
                price float)""")
    '''

    conn = sqlite3.connect("total_cash_file.db")
    c = conn.cursor()


    '''
    c.execute("""CREATE TABLE items (
                date_given integer,
                euros_given float)""")
    '''


    def open_total_balance():
        balance_w = Tk()
        balance_w.title("ΣΥΝΟΛΙΚΟ ΥΠΟΛΟΙΠΟ")
        balance_w.geometry("540x290")
        balance_w.config(background="#CAE1DA")

        def total_balance_euros():
            conn = sqlite3.connect("pesticide_and_fertilizers_file.db")
            c = conn.cursor()

            c.execute("SELECT price FROM products")
            rows = c.fetchall()
            total_euros = sum(
                float(str(row[0]).replace(',', '.')) for row in rows if
                row[0])

            conn = sqlite3.connect("total_cash_file.db")
            c = conn.cursor()

            c.execute("SELECT euros_given FROM items")
            rows = c.fetchall()
            total_euro = sum(
                float(str(row[0]).replace(',', '.')) for row in rows if
                row[0])

            total_balance = total_euros - total_euro

            balance_label.config(text="ΣΥΝΟΛΙΚΟ ΥΠΟΛΟΙΠΟ : {:.2f} €".format(total_balance))

            return total_balance


        balance_btn = Button(balance_w, text=" ΣΥΝΟΛΙΚΟ ΥΠΟΛΟΙΠΟ ", font=("arial black", 11), command=total_balance_euros, bd=6,
                           background="#06D784", activebackground="#06D784", activeforeground="#06D784")
        balance_btn.grid(row=1, pady=(40, 5), columnspan=2, padx=40, ipadx=40)
        balance_label = Label(balance_w, text=" ", font=("arial black", 13), width=30, background="#BAE5F1",
                                  relief="ridge", borderwidth=10)
        balance_label.grid(row=2, columnspan=2, pady=30, padx=20, ipadx=50)

    def submit():
        conn = sqlite3.connect("pesticide_and_fertilizers_file.db")
        c = conn.cursor()

        c.execute("INSERT INTO products VALUES (:date, :number, :items, :description, :price)",
                  {
                      "date": date.get(),
                      "number": number.get(),
                      "items": items.get(),
                      "description": description.get(),
                      "price": price.get()
                  })
        price_value = price.get()
        if price_value == "0" or not price_value.strip():
            price_value = ""


        confirmation = messagebox.askquestion("Εισαγωγή Αρχείων", "Do You Want To Insert this Files ???")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo(" Εισαγωγή Αρχείων ", "Επιτυχής Εισαγωγή Αρχείων. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Εισαγωγής", "Η Εισαγωγή Αρχείων ακυρώθηκε. !!!")
        conn.close()

        date.delete(0, END)
        number.delete(0, END)
        items.delete(0, END)
        description.delete(0, END)
        price.delete(0, END)

    def save_to_excel():
        conn = sqlite3.connect("pesticide_and_fertilizers_file.db")
        c = conn.cursor()

        c.execute("SELECT * FROM products")
        data = c.fetchall()

        try:
            workbook = load_workbook("pesticide_and_fertilizers.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            headers = ["ΗΜΕΡΟΜΗΝΙΑ", "ΑΡ.ΤΙΜΟΛΟΓΙΟΥ", "ΤΕΜΑΧΙΑ", "ΠΕΡΙΓΡΑΦΗ", "ΤΙΜΗ"]
            sheet.append(headers)

        # Clear existing data in the worksheet
        sheet.delete_rows(2, sheet.max_row)

        for row in data:
            sheet.append(row)

        workbook.save("pesticide_and_fertilizers.xlsx")

        conn.close()

        messagebox.showinfo("Αποθήκευση Δεδομένων", "Τα δεδομένα αποθηκεύτηκαν σε αρχείο Excel !!!")


    def delete_all_data():
        confirmation = messagebox.askyesno("Confirmation", "Είστε σίγουρος ότι θέλετε να διαγράψετε όλα τα Αρχεία ;")
        if confirmation:
            conn = sqlite3.connect("pesticide_and_fertilizers_file.db")
            c = conn.cursor()
            c.execute("DELETE FROM products")
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Ολα τα Αρχεία διαγράφηκαν με επιτυχία !!")


    def show():
        show = Tk()
        show.title("pesticide_and_fertilizers DATABASE")
        show.geometry("700x700")
        show.config(background="#CAE1DA")

        show_id = choose_id.get()
        conn = sqlite3.connect("pesticide_and_fertilizers_file.db")
        c = conn.cursor()

        c.execute("SELECT *,  oid FROM products")
        elements = c.fetchall()


        for i, element in enumerate(elements):
            text = '   '.join(str(item) for item in element)
            show_label = Label(show, text=text, font=("arial black", 10), relief="ridge", bd=7, borderwidth=5,
                               background="#D1F3F3", foreground="#20047A")
            show_label.grid(row=i, column=0, pady=1, sticky=W, ipadx=30)


        save_button = Button(show, text="Αποθήκευση σε 'Excel' ", command=save_to_excel, font=("arial black", 11),
                             background="#06D784")
        save_button.grid(row=i + 1, column=0, pady=10, ipadx=26, sticky=W, padx=30)
        delete_button = Button(show, text="Διαγραφή όλων των Αρχείων", command=delete_all_data, font=("arial black", 11),
                               background="#EA6969")
        delete_button.grid(row=i + 1, column=0, pady=10, padx=350, ipadx=8)


    def update():
        conn = sqlite3.connect("pesticide_and_fertilizers_file.db")
        c = conn.cursor()
        show_id = choose_id.get()

        c.execute("""UPDATE products SET
                   date = :date,
                   number = :number,
                   items = :items,
                   description = :description,
                   price = :price
    
    
                   WHERE oid = :oid""",
                  {'date': date_editor.get(),
                   'number': number_editor.get(),
                   'items': items_editor.get(),
                   'description': description_editor.get(),
                   'price': price_editor.get(),
                   'oid': show_id})

        confirmation = messagebox.askquestion("Επεξεργασία Αρχείων", "Do you want to save the changes  ???")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo(" Αποθήκευση Αρχείων ", "Επιτυχής αλλαγή και αποθήκευση Αρχείων. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Εισαγωγής", "Η αλλαγή Αρχείων ακυρώθηκε. !!!")

        conn.close()
        editor.destroy()

    def edit():
        global editor
        editor = Tk()
        editor.title("ΕΠΕΞΕΡΓΑΣΙΑ ΣΤΟΙΧΕΙΩΝ")
        editor.geometry("600x400")
        editor.config(background="#CAE1DA")

        conn = sqlite3.connect("pesticide_and_fertilizers_file.db")
        c = conn.cursor()
        show_id = choose_id.get()

        c.execute("SELECT * FROM products WHERE oid = " + show_id)
        shows = c.fetchall()

        global date_editor
        global number_editor
        global items_editor
        global description_editor
        global price_editor

        date_editor_label = Label(editor, text=" Ημερομηνία : ", font=("arial black", 12), background="#CAE1DA")
        date_editor_label.grid(row=1, column=0, sticky=W, padx=20, pady=(20, 10))
        number_editor_label = Label(editor, text=" Αρ. Τιμολογίου : ", font=("arial black", 12), background="#CAE1DA")
        number_editor_label.grid(row=2, column=0, sticky=W, padx=20, pady=10)
        items_editor_label = Label(editor, text=" Τεμάχια : ", font=("arial black", 12), background="#CAE1DA")
        items_editor_label.grid(row=3, column=0, sticky=W, padx=20, pady=10)
        description_editor_label = Label(editor, text=" Περιγραφή : ", font=("arial black", 12), background="#CAE1DA")
        description_editor_label.grid(row=4, column=0, sticky=W, padx=20, pady=10)
        price_editor_label = Label(editor, text=" Τιμή € : ", font=("arial black", 12), background="#CAE1DA")
        price_editor_label.grid(row=5, column=0, sticky=W, padx=20, pady=10)

        date_editor = Entry(editor, width=25, font=("arial black", 12), bd=3, background="#DBE5E5")
        date_editor.grid(row=1, column=1, padx=20, sticky=W, pady=(20, 10))
        number_editor = Entry(editor, width=25, font=("arial black", 12), bd=3, background="#DBE5E5")
        number_editor.grid(row=2, column=1, padx=20, sticky=W, pady=10)
        items_editor = Entry(editor, width=25, font=("arial black", 12), bd=3, background="#DBE5E5")
        items_editor.grid(row=3, column=1, padx=20, sticky=W, pady=10)
        description_editor = Entry(editor, width=25, font=("arial black", 12), bd=3, background="#DBE5E5")
        description_editor.grid(row=4, column=1, padx=20, sticky=W, pady=10)
        price_editor = Entry(editor, width=25, font=("arial black", 12), bd=3, background="#DBE5E5")
        price_editor.grid(row=5, column=1, padx=20, sticky=W, pady=10)

        for show in shows:
            date_editor.insert(0, show[0])
            number_editor.insert(0, show[1])
            items_editor.insert(0, show[2])
            description_editor.insert(0, show[3])
            price_editor.insert(0, show[4])


        edit_btn = Button(editor, text="Επεξεργασία και Αποθήκευση \n Στοιχείων", bd=4, command=update, font=("arial black", 10), background="#06D784", activeforeground="#06D784", activebackground="#06D784")
        edit_btn.grid(row=6, column=0, columnspan=2, pady=30, padx=10, ipadx=30)




    def delete():
        conn = sqlite3.connect("pesticide_and_fertilizers_file.db")
        c = conn.cursor()

        c.execute("DELETE from products WHERE oid = " + choose_id.get())

        confirmation = messagebox.askquestion("Διαγραφή Στοιχείων", " Θέλετε να διαγράψετε τα Στοιχεία ;")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo("Η Διαγραφή Ολοκληρώθηκε", "Τα στοιχεία διαγράφηκαν με επιτυχία. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Διαγραφής", "Η διαγραφή ακυρώθηκε. !!!")

        conn.commit()
        conn.close()



    def open_total_euros():
        global total_euros_label
        total_w = Tk()
        total_w.title("ΣΥΝΟΛΙΚΑ ΕΥΡΩ")
        total_w.geometry("540x270")
        total_w.config(background="#CAE1DA")



        def total_euros():

            conn = sqlite3.connect("pesticide_and_fertilizers_file.db")
            c = conn.cursor()

            c.execute("SELECT price FROM products")
            rows = c.fetchall()
            total_euros = sum(
                float(str(row[0]).replace(',', '.')) for row in rows if
                row[0])

            conn.close()
            total_euros_label.config(text="ΣΥΝΟΛΙΚΑ ΕΥΡΩ : {:.2f} €".format(total_euros))

            return total_euros



        total_btn = Button(total_w, text=" ΣΥΝΟΛΟ ΕΥΡΩ ", font=("arial black", 11), command=total_euros, bd=6,
                           background="#06D784", activebackground="#06D784", activeforeground="#06D784")
        total_btn.grid(row=1, pady=(40, 5), columnspan=2, padx=40, ipadx=65)
        total_euros_label = Label(total_w, text=" ", font=("arial black", 13), width=30, background="#BAE5F1",
                                  relief="ridge", borderwidth=10)
        total_euros_label.grid(row=2, columnspan=2, pady=30, padx=20, ipadx=50)

        return total_euros_label

    def total_cash():
        cash_w = Tk()
        cash_w.title("ΜΕΤΡΗΤΑ")
        cash_w.geometry("680x500")
        cash_w.config(background="#CAE1DA")



        def submit_cash():
            conn = sqlite3.connect("total_cash_file.db")
            c = conn.cursor()

            c.execute("INSERT INTO items VALUES (:date_given, :euros_given)",
                      {
                          "date_given": date_given.get(),
                          "euros_given": euros_given.get()
                      })
            euros_given_value = euros_given.get()
            if euros_given_value == "0" or not euros_given_value.strip():
                euros_given_value = ""

            confirmation = messagebox.askquestion("Εισαγωγή Αρχείων", "Do You Want To Insert this Files ???")
            if confirmation == 'yes':
                conn.commit()
                messagebox.showinfo(" Εισαγωγή Αρχείων ", "Επιτυχής Εισαγωγή Αρχείων. !!!")
            else:
                conn.rollback()
                messagebox.showinfo("Ακύρωση Εισαγωγής", "Η Εισαγωγή Αρχείων ακυρώθηκε. !!!")

            conn.close()

            date_given.delete(0, END)
            euros_given.delete(0, END)



        def save_cash_to_excel():
            conn = sqlite3.connect("total_cash_file.db")
            c = conn.cursor()

            c.execute("SELECT * FROM items")
            data = c.fetchall()

            try:
                workbook = load_workbook("total_cash.xlsx")
                sheet = workbook.active
            except FileNotFoundError:
                workbook = Workbook()
                sheet = workbook.active
                headers = ["ΗΜΕΡΟΜΗΝΙΑ", "ΜΕΤΡΗΤΑ"]
                sheet.append(headers)

            # Clear existing data in the worksheet
            sheet.delete_rows(2, sheet.max_row)

            for row in data:
                sheet.append(row)

            workbook.save("total_cash.xlsx")

            conn.close()

            messagebox.showinfo("Αποθήκευση Δεδομένων", "Τα δεδομένα αποθηκεύτηκαν σε αρχείο Excel !!!")

        def delete_all_data_cash():
            confirmation = messagebox.askyesno("Confirmation", "Είστε σίγουρος ότι θέλετε να διαγράψετε όλα τα Αρχεία ;")
            if confirmation:
                conn = sqlite3.connect("total_cash_file.db")
                c = conn.cursor()
                c.execute("DELETE FROM items")
                conn.commit()
                conn.close()
                messagebox.showinfo("Success", "Ολα τα Αρχεία διαγράφηκαν με επιτυχία !!")

        def show_cash():
            show_w = Tk()
            show_w.title("TOTAL CASH DATABASE")
            show_w.geometry("500x650")
            show_w.config(background="#CAE1DA")

            show_id = choose_csh_id.get()
            conn = sqlite3.connect("total_cash_file.db")
            c = conn.cursor()

            c.execute("SELECT *,  oid FROM items")
            elements = c.fetchall()

            for i, element in enumerate(elements):
                text = '          '.join(str(item) for item in element)
                show_label = Label(show_w, text=text, font=("arial black", 10), relief="ridge", bd=7, borderwidth=5,
                                   background="#D1F3F3", foreground="#20047A")
                show_label.grid(row=i, column=0, pady=1, sticky=W, ipadx=30)

            save_button = Button(show_w, text="Αποθήκευση σε \n'Excel' ", command=save_cash_to_excel, font=("arial black", 10),
                                 background="#06D784")
            save_button.grid(row=i + 1, column=0, pady=5, ipadx=26, sticky=W, padx=10)
            delete_button = Button(show_w, text="Διαγραφή όλων \nτων Αρχείων", command=delete_all_data_cash,
                                   font=("arial black", 10),
                                   background="#EA6969")
            delete_button.grid(row=i + 1, column=1, pady=5, sticky=W, padx=10, ipadx=6)

        def delete_cash():
            conn = sqlite3.connect("total_cash_file.db")
            c = conn.cursor()

            c.execute("DELETE from items WHERE oid = " + choose_csh_id.get())

            confirmation = messagebox.askquestion("Διαγραφή Στοιχείων", " Θέλετε να διαγράψετε τα Στοιχεία ;")
            if confirmation == 'yes':
                conn.commit()
                messagebox.showinfo("Η Διαγραφή Ολοκληρώθηκε", "Τα στοιχεία διαγράφηκαν με επιτυχία. !!!")
            else:
                conn.rollback()
                messagebox.showinfo("Ακύρωση Διαγραφής", "Η διαγραφή ακυρώθηκε. !!!")

            conn.commit()
            conn.close()


        def update_cash():
            conn = sqlite3.connect("total_cash_file.db")
            c = conn.cursor()
            show_id = choose_csh_id.get()

            c.execute("""UPDATE items SET
                           date_given = :date_given,
                           euros_given = :euros_given
                           
                           WHERE oid = :oid""",
                          {'date_given': date_given_editor.get(),
                           'euros_given': euros_given_editor.get(),
                           'oid': show_id})

            confirmation = messagebox.askquestion("Επεξεργασία Αρχείων", "Do you want to save the changes  ???")
            if confirmation == 'yes':
                conn.commit()
                messagebox.showinfo(" Αποθήκευση Αρχείων ", "Επιτυχής αλλαγή και αποθήκευση Αρχείων. !!!")
            else:
                conn.rollback()
                messagebox.showinfo("Ακύρωση Εισαγωγής", "Η αλλαγή Αρχείων ακυρώθηκε. !!!")

            conn.close()
            editor_csh.destroy()

        def edit_cash():
            global editor_csh
            editor_csh = Tk()
            editor_csh.title("ΕΠΕΞΕΡΓΑΣΙΑ ΣΤΟΙΧΕΙΩΝ")
            editor_csh.geometry("600x400")
            editor_csh.config(background="#CAE1DA")

            conn = sqlite3.connect("total_cash_file.db")
            c = conn.cursor()
            show_id = choose_csh_id.get()

            c.execute("SELECT * FROM items WHERE oid = " + show_id)
            shows = c.fetchall()

            global date_given_editor
            global euros_given_editor

            date_given_editor_label = Label(editor_csh, text="    ΗΜΕΡ/ΝΙΑ :   ", font=("arial black", 14), relief="ridge", bd=5,
                                     background="#DC9127")
            date_given_editor_label.grid(row=1, column=0, sticky=W, padx=20, pady=(20, 10))
            euros_given_editor_label = Label(editor_csh, text="    ΜΕΤΡΗΤΑ :   ", font=("arial black", 14), relief="ridge", bd=5,
                                      background="#DC9127")
            euros_given_editor_label.grid(row=2, column=0, sticky=W, padx=20, pady=(20, 10))

            date_given_editor = Entry(editor_csh, width=15, font=("arial black", 14), bd=3, background="#DBE5E5")
            date_given_editor.grid(row=1, column=1, padx=80, sticky=W, pady=(20, 10))
            euros_given_editor = Entry(editor_csh, width=15, font=("arial black", 14), bd=3, background="#DBE5E5")
            euros_given_editor.grid(row=2, column=1, padx=80, sticky=W, pady=(20, 10))




            for show in shows:
                date_given_editor.insert(0, show[0])
                euros_given_editor.insert(0, show[1])


            edit_btn = Button(editor_csh, text="Επεξεργασία και Αποθήκευση \n Στοιχείων", bd=4, command=update_cash,
                              font=("arial black", 10), background="#06D784", activeforeground="#06D784",
                              activebackground="#06D784")
            edit_btn.grid(row=6, column=0, columnspan=2, pady=30, padx=10, ipadx=30)


        def total_euros_given_cash():
            conn = sqlite3.connect("total_cash_file.db")
            c = conn.cursor()

            c.execute("SELECT euros_given FROM items")
            rows = c.fetchall()
            total_euro = sum(
                float(str(row[0]).replace(',', '.')) for row in rows if
                row[0])

            conn.close()
            total_euros_given_label.config(text="ΣΥΝΟΛΙΚΑ ΜΕΤΡΗΤΑ : {:.2f} €".format(total_euro))


            return total_euro






        date_given_label = Label(cash_w, text="    ΗΜΕΡ/ΝΙΑ :   ", font=("arial black", 12), relief="ridge", bd=5, background="#DC9127")
        date_given_label.grid(row=1, column=0, sticky=W, padx=20, pady=(20, 10))
        euros_given_label = Label(cash_w, text="    ΜΕΤΡΗΤΑ :   ", font=("arial black", 12), relief="ridge", bd=5, background="#DC9127")
        euros_given_label.grid(row=2, column=0, sticky=W, padx=20, pady=(20, 10))
        choose_csh_id_label = Label(cash_w, text="    ΕΠΙΛΟΓΗ ID :   ", font=("arial black", 12), relief="ridge", bd=5, background="#DC9127")
        choose_csh_id_label.grid(row=3, column=0, sticky=W, padx=20, pady=(20, 10))

        date_given = Entry(cash_w, width=15, font=("arial black", 12), bd=3, background="#DBE5E5")
        date_given.grid(row=1, column=1, padx=80, sticky=W, pady=(20, 10))
        euros_given = Entry(cash_w, width=15, font=("arial black", 12), bd=3, background="#DBE5E5")
        euros_given.grid(row=2, column=1, padx=80, sticky=W, pady=(20, 10))
        choose_csh_id = Entry(cash_w, width=8, font=("arial black", 12), bd=3, background="#DBE5E5")
        choose_csh_id.grid(row=3, column=1, padx=80, sticky=W, pady=(20, 10))




        total_euros_given_label = Label(cash_w, text=" ", font=("arial black", 13), width=35, background="#BAE5F1",
                                    relief="ridge", borderwidth=10)
        total_euros_given_label.grid(row=9, columnspan=2, pady=15, padx=40, ipadx=50)

        total_euros_given_btn = Button(cash_w, text=" ΣΥΝΟΛΙΚΑ ΜΕΤΡΗΤΑ ", font=("arial black", 11),
                                   command=total_euros_given_cash, bd=6,
                                   background="#06D784", activebackground="#06D784", activeforeground="#06D784")
        total_euros_given_btn.grid(row=8, columnspan=2, pady=(20, 5), ipadx=30)




        submit_btn = Button(cash_w, text="Εισαγωγή Στοιχείων", command=submit_cash, font=("arial black", 10), bd=4,
                            background="#06D784", activebackground="#06D784", activeforeground="#06D784")
        submit_btn.grid(row=6, column=0, pady=(20, 10), padx=10, ipadx=40)

        show_btn = Button(cash_w, text="Εμφάνιση Στοιχείων", command=show_cash, font=("arial black", 10), bd=4,
                          background="#0499A0", activebackground="#0499A0", activeforeground="#0499A0")
        show_btn.grid(row=7, column=1, pady=10, padx=10, ipadx=40)

        delete_btn = Button(cash_w, text="Διαγραφή Στοιχείων", command=delete_cash, font=("arial black", 10), bd=4,
                            background="#F67EA3", activebackground="#F67EA3", activeforeground="#F67EA3")
        delete_btn.grid(row=7, column=0, pady=10, padx=10, ipadx=40)

        edit_btn = Button(cash_w, text="Επεξεργασία Στοιχείων", command=edit_cash, font=("arial black", 10), bd=4,
                          background="#A481C9", activebackground="#A481C9", activeforeground="#A481C9")
        edit_btn.grid(row=6, column=1, pady=(20, 10), padx=10, ipadx=28)





    date_label = Label(root, text=" Ημερομηνία : ", font=("arial black", 12), background="#CAE1DA")
    date_label.grid(row=1, column=0, sticky=W, padx=20, pady=(20, 10))
    number_label = Label(root, text=" Αρ. Τιμολογίου : ", font=("arial black", 12), background="#CAE1DA")
    number_label.grid(row=2, column=0, sticky=W, padx=20, pady=10)
    items_label = Label(root, text=" Τεμάχια : ", font=("arial black", 12), background="#CAE1DA")
    items_label.grid(row=3, column=0, sticky=W, padx=20, pady=10)
    description_label = Label(root, text=" Περιγραφή : ", font=("arial black", 12), background="#CAE1DA")
    description_label.grid(row=4, column=0, sticky=W, padx=20, pady=10)
    price_label = Label(root, text=" Τιμή € : ", font=("arial black", 12), background="#CAE1DA")
    price_label.grid(row=5, column=0, sticky=W, padx=20, pady=10)
    choose_id_label = Label(root, text=" Επιλογή ID : ", font=("arial black", 12), background="#CAE1DA")
    choose_id_label.grid(row=6, column=0, sticky=W, padx=20, pady=10)



    date = Entry(root, width=25, font=("arial black", 12), bd=3, background="#DBE5E5")
    date.grid(row=1, column=1, padx=20, sticky=W, pady=(20, 10))
    number = Entry(root, width=25, font=("arial black", 12), bd=3, background="#DBE5E5")
    number.grid(row=2, column=1, padx=20, sticky=W, pady=10)
    items = Entry(root, width=25, font=("arial black", 12), bd=3, background="#DBE5E5")
    items.grid(row=3, column=1, padx=20, sticky=W, pady=10)
    description = Entry(root, width=25, font=("arial black", 12), bd=3, background="#DBE5E5")
    description.grid(row=4, column=1, padx=20, sticky=W, pady=10)
    price = Entry(root, width=25, font=("arial black", 12), bd=3, background="#DBE5E5")
    price.grid(row=5, column=1, padx=20, sticky=W, pady=10)
    choose_id = Entry(root, width=7, font=("arial black", 12), bd=3, background="#DBE5E5")
    choose_id.grid(row=6, column=1, padx=20, sticky=W, pady=10)


    submit_btn = Button(root, text="Εισαγωγή Στοιχείων", command=submit, font=("arial black", 12), bd=4, background="#06D784", activebackground="#06D784", activeforeground="#06D784")
    submit_btn.grid(row=7, column=0, pady=(20, 10), padx=10, ipadx=40)

    show_btn = Button(root, text="Εμφάνιση Στοιχείων", command=show, font=("arial black", 12), bd=4,  background="#0499A0", activebackground="#0499A0", activeforeground="#0499A0")
    show_btn.grid(row=8, column=1,  pady=10, padx=10, ipadx=40)

    delete_btn = Button(root, text="Διαγραφή Στοιχείων", command=delete, font=("arial black", 12), bd=4, background="#F67EA3", activebackground="#F67EA3", activeforeground="#F67EA3")
    delete_btn.grid(row=8, column=0, pady=10, padx=10, ipadx=40)

    edit_btn = Button(root, text="Επεξεργασία Στοιχείων", command=edit, font=("arial black", 12), bd=4, background="#A481C9", activebackground="#A481C9", activeforeground="#A481C9")
    edit_btn.grid(row=7, column=1, pady=(20, 10), padx=10, ipadx=28)


    cash_btn = Button(root, text=" ΜΕΤΡΗΤΑ € ",  font=("arial black", 11), command=total_cash,  bd=6, background="#08CBCB", activebackground="#08CBCB", activeforeground="#08CBCB")
    cash_btn.grid(row=9, column=1, pady=(20, 5), ipadx=80)
    total_btn = Button(root, text=" ΣΥΝΟΛΟ € ",  font=("arial black", 11), command=open_total_euros,  bd=6, background="#08CBCB", activebackground="#08CBCB", activeforeground="#08CBCB")
    total_btn.grid(row=9, column=0, pady=(20, 5), ipadx=80)

    total_balance_btn = Button(root, text="  ΥΠΟΛΟΙΠΟ €  ",  font=("arial black", 11), command=open_total_balance,  bd=6, background="#08CBCB", activebackground="#08CBCB", activeforeground="#08CBCB")
    total_balance_btn.grid(row=10, column=0, columnspan=2, pady=(20, 5), ipadx=70)



    name_label = Label(root, text="Created and Designed by : Papaioannou Antonios", font=("arial black", 11), foreground="grey", background="#CAE1DA", borderwidth=1)
    name_label.grid(columnspan=2, row=18, sticky=E, pady=(5, 0))




    conn.commit()
    conn.close()

    root.mainloop()